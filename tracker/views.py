from django.shortcuts import render, redirect
from django.shortcuts import get_object_or_404
from openpyxl import Workbook
from django.http import HttpResponse
from django.contrib.auth import authenticate, login, logout
from django.contrib.auth.decorators import login_required
from django.utils import timezone
from django.db.models import Count
from .models import User
from .models import Asset, Department, Maintenance


def login_view(request):
    if request.method == 'POST':
        username = request.POST['username']
        password = request.POST['password']
        user = authenticate(request, username=username, password=password)
        if user and user.is_staff:
            login(request, user)
            if user.role == 'admin':
                return redirect('admin_dashboard')
            elif user.role == 'hod':
                return redirect('hod_dashboard')
            else:
                return redirect('staff_dashboard')
        else:
            return render(request, 'tracker/login.html', {'error': 'Invalid credentials or not allowed'})
    return render(request, 'tracker/login.html')


def logout_view(request):
    logout(request)
    return redirect('login')


def home_redirect(request):
    return redirect('login')

@login_required
def admin_dashboard(request):
    if request.user.role != 'admin':
        return redirect('login')

    from django.db.models import Count, Sum

    assets = Asset.objects.all()
    maintenance = Maintenance.objects.all()
    users = User.objects.all()
    departments = Department.objects.all()

    total_assets = assets.count()
    total_users = users.count()
    total_departments = departments.count()
    total_maintenance = maintenance.count()
    total_maintenance_cost = maintenance.aggregate(Sum('cost'))['cost__sum'] or 0

    chart_data = assets.values('status').annotate(count=Count('id'))
    status_data = [0, 0, 0, 0]
    for item in chart_data:
        if item['status'] == 'available': status_data[0] = item['count']
        elif item['status'] == 'in_use': status_data[1] = item['count']
        elif item['status'] == 'maintenance': status_data[2] = item['count']
        elif item['status'] == 'retired': status_data[3] = item['count']

    return render(request, 'tracker/admin_dashboard.html', {
        'total_assets': total_assets,
        'total_users': total_users,
        'total_departments': total_departments,
        'total_maintenance': total_maintenance,
        'total_maintenance_cost': total_maintenance_cost,
        'chart_data': status_data,
        'assets': assets,
        'maintenance': maintenance,
        'users': users,
        'departments': departments,
    })

@login_required
def hod_dashboard(request):
    if request.user.role != 'hod':
        return redirect('login')

    dept = request.user.department
    department_assets = Asset.objects.filter(department=dept).count()
    department_users = User.objects.filter(department=dept).count()
    department_maintenance = Maintenance.objects.filter(asset__department=dept).count()

    chart_data = [0, 0, 0, 0]
    for item in Asset.objects.filter(department=dept).values('status').annotate(count=Count('id')):
        if item['status'] == 'available':
            chart_data[0] = item['count']
        elif item['status'] == 'in_use':
            chart_data[1] = item['count']
        elif item['status'] == 'maintenance':
            chart_data[2] = item['count']
        elif item['status'] == 'retired':
            chart_data[3] = item['count']

    context = {
        'department_assets': department_assets,
        'department_users': department_users,
        'department_maintenance': department_maintenance,
        'chart_data': chart_data,
    }
    return render(request, 'tracker/hod_dashboard.html', context)





@login_required
def staff_dashboard(request):
    if request.user.role != 'staff':
        return redirect('login')

    assets = Asset.objects.filter(assigned_to=request.user)
    total_assets = assets.count()
    total_maintenance = Maintenance.objects.filter(asset__assigned_to=request.user).count()

    chart_data = assets.values('status').annotate(count=Count('id'))
    status_data = [0, 0, 0, 0]
    for item in chart_data:
        if item['status'] == 'available':
            status_data[0] = item['count']
        elif item['status'] == 'in_use':
            status_data[1] = item['count']
        elif item['status'] == 'maintenance':
            status_data[2] = item['count']
        elif item['status'] == 'retired':
            status_data[3] = item['count']

    return render(request, 'tracker/staff_dashboard.html', {
        'total_assets': total_assets,
        'total_maintenance': total_maintenance,
        'chart_data': status_data,
    })

@login_required
def view_assets(request):
    user = request.user

    if user.role == 'admin':
        assets = Asset.objects.all()
    elif user.role == 'hod':
        assets = Asset.objects.filter(department=user.department)
    else:  # staff
        assets = Asset.objects.filter(assigned_to=user)

    return render(request, 'tracker/assets_list.html', {'assets': assets})
@login_required
def log_maintenance(request, asset_id):
    asset = get_object_or_404(Asset, id=asset_id)

    # Ensure only assigned staff can log maintenance
    if request.user.role == 'staff' and asset.assigned_to != request.user:
        return redirect('staff_dashboard')

    if request.method == 'POST':
        description = request.POST['description']
        cost = request.POST['cost']
        date = request.POST.get('maintenance_date') or timezone.now().date()

        Maintenance.objects.create(
            asset=asset,
            maintenance_date=date,
            description=description,
            cost=cost
        )
        return redirect('view_assets')

    return render(request, 'tracker/log_maintenance.html', {'asset': asset})
@login_required
def view_maintenance(request):
    user = request.user

    if user.role == 'admin':
        records = Maintenance.objects.all()
    elif user.role == 'hod':
        records = Maintenance.objects.filter(asset__department=user.department)
    else:
        records = Maintenance.objects.filter(asset__assigned_to=user)

    return render(request, 'tracker/maintenance_list.html', {'records': records})

# views.py – Final General Excel Report



@login_required
def export_overall_report(request):
    user = request.user
    wb = Workbook()

    # 1. Summary Sheet (Admin only)
    if user.role == 'admin':
        ws_summary = wb.active
        ws_summary.title = "Summary"
        ws_summary.append(['Metric', 'Count'])
        ws_summary.append(['Total Assets', Asset.objects.count()])
        ws_summary.append(['Total Maintenance Logs', Maintenance.objects.count()])
        ws_summary.append(['Total Users', User.objects.count()])
        ws_summary.append(['Total Departments', Department.objects.count()])
    else:
        wb.remove(wb.active)

    # 2. Assets Sheet
    ws_assets = wb.create_sheet("Assets")
    ws_assets.append(['Serial No', 'Name', 'Status', 'Department', 'Assigned To'])
    if user.role == 'admin':
        assets = Asset.objects.all()
    elif user.role == 'hod':
        assets = Asset.objects.filter(department=user.department)
    else:
        assets = Asset.objects.filter(assigned_to=user)
    for asset in assets:
        ws_assets.append([
            asset.serial_number,
            asset.asset_name,
            asset.get_status_display(),
            asset.department.name if asset.department else '',
            asset.assigned_to.username if asset.assigned_to else ''
        ])

    # 3. Maintenance Sheet
    ws_maint = wb.create_sheet("Maintenance")
    ws_maint.append(['Asset', 'Date', 'Description', 'Cost'])
    if user.role == 'admin':
        maints = Maintenance.objects.all()
    elif user.role == 'hod':
        maints = Maintenance.objects.filter(asset__department=user.department)
    else:
        maints = Maintenance.objects.filter(asset__assigned_to=user)
    for m in maints:
        ws_maint.append([
            m.asset.asset_name,
            m.maintenance_date.strftime('%Y-%m-%d'),
            m.description,
            float(m.cost),
        ])

    # 4. Users Sheet (Admin only)
    if user.role == 'admin':
        ws_users = wb.create_sheet("Users")
        ws_users.append(['Username', 'Role', 'Department'])
        for u in User.objects.all():
            ws_users.append([u.username, u.role, u.department.name if u.department else ''])

    # 5. Departments Sheet (Admin only)
    if user.role == 'admin':
        ws_depts = wb.create_sheet("Departments")
        ws_depts.append(['Name', 'Description'])
        for d in Department.objects.all():
            ws_depts.append([d.name, d.description])

    # Return Excel file
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=overall_report.xlsx'
    wb.save(response)
    return response
